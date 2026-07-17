import io
from collections import defaultdict
from collections.abc import Mapping, Sequence
from datetime import UTC, date, datetime
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

LEADS_XLSX_COLUMNS: list[tuple[str, str]] = [
    ("username", "Юзернейм"),
    ("name", "Имя"),
    ("telegram_name", "Имя в Telegram"),
    ("niche", "Ниша"),
    ("main_problem", "Основная проблема"),
    ("average_check", "Средний чек"),
    ("revenue_estimate", "Заработок"),
    ("sales_volume", "Количество продаж"),
    ("lead_temperature", "Температура лида"),
    ("confidence", "Уверенность анализа"),
    ("summary", "Краткое резюме"),
]

_MOSCOW_TZ = ZoneInfo("Europe/Moscow")
_MAX_COLUMN_WIDTH = 60


def _excel_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime | date):
        return value.isoformat()
    return str(value)


def _lead_datetime(value: Any) -> datetime | None:
    if not isinstance(value, datetime):
        return None
    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC)
    return value.astimezone(UTC)


def _set_text(cell: Cell, value: str) -> None:
    cell.value = value
    cell.data_type = "s"
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def _populate_sheet(
    workbook: Workbook,
    title: str,
    rows: Sequence[Mapping[str, Any]],
    table_name: str | None,
) -> None:
    worksheet = workbook.create_sheet(title=title)
    worksheet.freeze_panes = "A2"

    headers = [header for _, header in LEADS_XLSX_COLUMNS]
    for column_index, header in enumerate(headers, start=1):
        _set_text(worksheet.cell(row=1, column=column_index), header)

    for row_index, row in enumerate(rows, start=2):
        for column_index, (key, _header) in enumerate(LEADS_XLSX_COLUMNS, start=1):
            _set_text(
                worksheet.cell(row=row_index, column=column_index),
                _excel_value(row.get(key)),
            )

    for column_index, header in enumerate(headers, start=1):
        values = [
            header,
            *(_excel_value(row.get(LEADS_XLSX_COLUMNS[column_index - 1][0])) for row in rows),
        ]
        width = min(max(len(value) for value in values) + 2, _MAX_COLUMN_WIDTH)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    if table_name is not None:
        last_column = get_column_letter(len(LEADS_XLSX_COLUMNS))
        table = Table(displayName=table_name, ref=f"A1:{last_column}{len(rows) + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)


def render_leads_xlsx(rows: Sequence[Mapping[str, Any]]) -> bytes:
    dated_rows: dict[date, list[Mapping[str, Any]]] = defaultdict(list)
    undated_rows: list[Mapping[str, Any]] = []

    for row in rows:
        lead_datetime = _lead_datetime(row.get("created_at"))
        if lead_datetime is None:
            undated_rows.append(row)
            continue
        lead_day = lead_datetime.astimezone(_MOSCOW_TZ).date()
        dated_rows[lead_day].append(row)

    workbook = Workbook()
    workbook.remove(workbook.active)

    if not dated_rows and not undated_rows:
        _populate_sheet(workbook, "Лиды", [], None)
    else:
        for lead_day in sorted(dated_rows, reverse=True):
            daily_rows = sorted(
                dated_rows[lead_day],
                key=lambda row: (
                    _lead_datetime(row.get("created_at")) or datetime.min.replace(tzinfo=UTC)
                ),
                reverse=True,
            )
            _populate_sheet(
                workbook,
                lead_day.strftime("%d.%m.%Y"),
                daily_rows,
                f"Leads_{lead_day.strftime('%Y%m%d')}",
            )
        if undated_rows:
            _populate_sheet(workbook, "Без даты", undated_rows, "Leads_undated")

    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()
