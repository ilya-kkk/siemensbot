from datetime import UTC, datetime
from io import BytesIO

from openpyxl import load_workbook

from app.services.leads import LEADS_XLSX_COLUMNS, render_leads_xlsx


def _load_export(rows: list[dict[str, object]]):
    return load_workbook(BytesIO(render_leads_xlsx(rows)))


def test_render_leads_xlsx_groups_by_moscow_day_and_sorts_newest_first() -> None:
    workbook = _load_export(
        [
            {
                "created_at": datetime(2026, 7, 1, 20, 30, tzinfo=UTC),
                "username": "first-day",
                "name": "Анна",
            },
            {
                "created_at": datetime(2026, 7, 2, 8, 0, tzinfo=UTC),
                "username": "newer",
                "name": "Борис",
            },
            {
                "created_at": datetime(2026, 7, 1, 21, 30, tzinfo=UTC),
                "username": "older",
                "name": "Вера",
            },
        ]
    )

    assert workbook.sheetnames == ["02.07.2026", "01.07.2026"]
    assert [cell.value for cell in workbook["02.07.2026"][1]] == [
        header for _, header in LEADS_XLSX_COLUMNS
    ]
    assert workbook["02.07.2026"]["A2"].value == "newer"
    assert workbook["02.07.2026"]["A3"].value == "older"
    assert workbook["01.07.2026"]["A2"].value == "first-day"


def test_render_leads_xlsx_formats_daily_sheet_as_excel_table() -> None:
    workbook = _load_export(
        [
            {
                "created_at": datetime(2026, 7, 1, 12, 30, tzinfo=UTC),
                "username": "client",
                "name": "Анна",
                "telegram_name": "Анна Иванова",
                "niche": "услуги",
                "main_problem": None,
                "average_check": "15000",
                "revenue_estimate": "300000",
                "sales_volume": "20",
                "summary": "Длинное резюме " * 20,
            }
        ]
    )
    worksheet = workbook["01.07.2026"]

    assert worksheet.freeze_panes == "A2"
    assert list(worksheet.tables) == ["Leads_20260701"]
    assert worksheet.tables["Leads_20260701"].ref == "A1:K2"
    assert worksheet.tables["Leads_20260701"].tableStyleInfo.name == "TableStyleMedium2"
    assert worksheet["E2"].value is None
    assert worksheet["K2"].alignment.wrap_text is True
    assert worksheet.column_dimensions["K"].width == 60


def test_render_leads_xlsx_writes_formula_like_values_as_text() -> None:
    formula_like_values = ["=1+1", "+SUM(A1:A2)", "-2+3", "@command"]
    workbook = _load_export(
        [
            {
                "created_at": datetime(2026, 7, 1, 12, 30, tzinfo=UTC),
                "username": value,
            }
            for value in formula_like_values
        ]
    )
    worksheet = workbook["01.07.2026"]

    assert [
        worksheet.cell(row=index, column=1).value for index in range(2, 6)
    ] == formula_like_values
    assert [worksheet.cell(row=index, column=1).data_type for index in range(2, 6)] == [
        "s",
        "s",
        "s",
        "s",
    ]


def test_render_leads_xlsx_keeps_rows_without_a_timestamp() -> None:
    workbook = _load_export(
        [
            {"created_at": datetime(2026, 7, 1, 12, 30, tzinfo=UTC), "username": "dated"},
            {"created_at": None, "username": "missing"},
            {"created_at": "not-a-date", "username": "invalid"},
        ]
    )

    assert workbook.sheetnames == ["01.07.2026", "Без даты"]
    assert workbook["Без даты"]["A2"].value == "missing"
    assert workbook["Без даты"]["A3"].value == "invalid"
    assert list(workbook["Без даты"].tables) == ["Leads_undated"]


def test_render_leads_xlsx_returns_header_only_workbook_when_empty() -> None:
    workbook = _load_export([])
    worksheet = workbook["Лиды"]

    assert workbook.sheetnames == ["Лиды"]
    assert worksheet.max_row == 1
    assert [cell.value for cell in worksheet[1]] == [header for _, header in LEADS_XLSX_COLUMNS]
    assert not worksheet.tables
